from selenium import webdriver as wd
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

from time import sleep
import pytest
import csv
from openpyxl import load_workbook as lw
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension as cd
from test_pyxl_lib import pyxl
import get_feature_bugs as gtb



@pytest.fixture
def driver(request):
    chrome_options = Options()
    chrome_options.add_argument("--window-size=1920,1080")
    wd = webdriver.Chrome(chrome_options=chrome_options)

    print(wd.capabilities)
    request.addfinalizer(wd.quit)
    return wd

# вызов функции для отладки выгрузки задач в разработке
def test_main(driver):
    # Создаем пустой ексель с
    df = gtb.cr_file_xls('В разработке.xlsx')

    # логинимся в системе.
    gtb.login(driver)

    # Найти все задачи по фильтру В разработке.
    iss_lst = gtb.get_tasks_list(driver, '10765')

    # Находим данные по каждой задаче и записываем их в итоговую таблицу
    count = 0

    for x in range(len(iss_lst)):
        iss = dev_tsk_data(driver, iss_lst[x])

        # Вызываем функцию записи в файл.
        count = write_to_xls(iss, df, count)


def dev_tsk_data(driver, issue, mode='proj_status'):
    task_res = []
    #print(issue)

    wait = WebDriverWait(driver, 5)
    driver.get(f'http://jira.it2g.ru/browse/{issue}')
    wait.until(EC.presence_of_element_located((By.ID, 'summary-val')))


    # Тема
    task_res.append(str(driver.find_element_by_id('summary-val').text))

    # Номер в Jira
    task_res.append(issue)

    if mode == 'proj_status':
        # Приоритет
        task_res.append(str(driver.find_element_by_id('priority-val').text))
        # Исполнитель
        task_res.append(str(driver.find_element_by_id('assignee-val').text))
    # Тип задачи
    # task_res.append(str(driver.find_element_by_id('type-val').text))

    else:
        # Статус
        task_res.append(str(driver.find_element_by_id('status-val').text))
        # Приоритет
        task_res.append(str(driver.find_element_by_id('priority-val').text))
        # Исправить в версии
        task_res.append(str(driver.find_element_by_id('fixfor-val').text))

    print(task_res)
    return task_res


# Запись в файл эксель данных по задаче и по связанным багам
def write_to_xls(task, df, page_name, lr=0, mode='proj_status'):
    border = Border(left = Side(border_style = 'thin', color = 'FF000000'),
                    right = Side(border_style = 'thin', color = 'FF000000'),
                    top = Side(border_style = 'thin', color = 'FF000000'),
                    bottom = Side(border_style = 'thin', color = 'FF000000'))

    font = Font(bold=True)

    wb = lw(df)
    ws1 = wb[page_name]
    starts, ends = '', ''

    if mode == 'bugs':
        headers = [
            'Тема:',
            '№ задачи Jira',
            'Статус',
            'Приоритет',
            'Релиз'
        ]



    if mode == 'proj_status':
        headers = [
            'Тема:',
            '№ задачи Jira',
            'Приоритет',

            'Исполнитель',
            'Комментарий'
        ]


    # assign the icon set to a rule



    for row in range(lr+1, lr+2):
        # Запишем заголовки в файл
        starts = f'{get_column_letter(1)}{lr+1}'
        if lr == 0:
            for col in range(0, len(headers)):
                _ = ws1.cell(column=col+1, row=row, value=headers[col])
                starts = f'{get_column_letter(col+1)}{lr+1}'
                ends = f'{get_column_letter(len(headers))}{row}'
                rang = f'{starts}:{ends}'
                #print(rang)
                gtb.style_range(ws1, rang, border=border, font=font)
                wb.save(filename = df)

        if type(task) is str:
            print(task)
            _ = ws1.cell(column=1, row=row+1, value=task)
            starts = f'{get_column_letter(1)}{row+1}'
            ends = f'{get_column_letter(len(headers))}{row+1}'
            rang = f'{starts}:{ends}'
            print(rang)
            gtb.style_range(ws1, rang, border=border, font=font)

            ws1.merge_cells(rang)
            wb.save(filename = df)
        else:
            if mode=='proj_status':
                for col in range(0, len(headers)-1):
                    if len(task) == 0:
                        _ = ws1.cell(column=col+1, row=row+1, value='')
                    else:
                        _ = ws1.cell(column=col+1, row=row+1, value=task[col])
                    ends = f'{get_column_letter(len(headers))}{row+1}'
                    rang = f'{starts}:{ends}'
                    #print(rang)
                    gtb.style_range(ws1, rang, border=border)
                    wb.save(filename = df)
            else:
                for col in range(0, len(headers)):
                    if len(task) == 0:
                        _ = ws1.cell(column=col+1, row=row+1, value='')
                    else:
                        _ = ws1.cell(column=col+1, row=row+1, value=task[col])
                    ends = f'{get_column_letter(len(headers))}{row+1}'
                    rang = f'{starts}:{ends}'
                    #print(rang)
                    gtb.style_range(ws1, rang, border=border)
                    wb.save(filename = df)
            #print(f'Финальная ячейка: {ends}')

        lr = row
    # Автоматическая ширина колонок.
    dims = {}
    for row in ws1.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws1.column_dimensions[col].width = value
    wb.save(filename = df)
    print('Загадочное lr ' + str(lr))
    return lr

