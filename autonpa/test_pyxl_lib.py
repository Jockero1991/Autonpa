from openpyxl import load_workbook as lw
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import csv
from openpyxl.formatting.rule import IconSet, FormatObject
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

def pyxl(p, sh, df):
    d_file = 'data\\' + df
    report_book = lw(d_file)
    current_sheet = report_book[sh]
    
    reading_txt=[]
    with open(p, "r") as csv_file:
        reader = csv.DictReader(csv_file, delimiter=';')
        #print(reader[2])
        for line in reader:
            reading_txt.append(line["№ в Jira|Тип задачи|Статус|Приоритет|Тема|Исполнитель|Тестировщик|Sprint"])
    #print(len(reading_txt))
    for row in range(3, (len(reading_txt)+3)):
        _ = current_sheet.cell(column = 2, row=row, value="{0}".format(reading_txt[row-3]))
    report_book.save(filename=d_file)
    

def example_pyxl():

    first = FormatObject(type='percent', val=0)
    second = FormatObject(type='percent', val=33)
    third = FormatObject(type='percent', val=67)
    iconset = IconSet(iconSet='3TrafficLights1', cfvo=[first, second, third], showValue=None, percent=None, reverse=None)
    # assign the icon set to a rule

    rule = Rule(type='iconSet', iconSet=iconset)
    wb = Workbook()
    dest_filename = 'empty_book.xlsx'

    ws1 = wb.active
    ws1.title = "range names"

    for row in range(1, 40):
        ws1.append(range(600))

    ws2 = wb.create_sheet(title="Pi")

    ws2['F5'] = 3.14

    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
            test = f'{get_column_letter(col)}{col}'
            print(f'Получаем первую ячейку заголовка: {test}')
    print(ws3['AA10'].value)
    ws1.conditional_formatting.add(f'{test}', rule)
    wb.save(filename = dest_filename)
#example_pyxl()
#test_pyxl('в_тестировании.csv', 'В тестировании')


def formatting_xls():
    wb = Workbook()
    ws = wb.active

    # Create fill
    redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
    # Add a two-color scale
    # Takes colors in excel 'RRGGBB' style.
    ws.conditional_formatting.add('A1:A10', ColorScaleRule(start_type='min', start_color='AA0000', end_type='max', end_color='00AA00'))

    # Add a three-color scale
    ws.conditional_formatting.add('B1:B10', ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000', mid_type='percentile', mid_value=50, mid_color='0000AA', end_type='percentile', end_value=90, end_color='00AA00'))

    # Add a conditional formatting based on a cell comparison
    # addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
    #  Format if cell is less than 'formula'
    ws.conditional_formatting.add('C2:C10', CellIsRule(operator='lessThan', formula=['C$1'], stopIfTrue=True, fill=redFill))

    # Format if cell is between 'formula'
    ws.conditional_formatting.add('D2:D10', CellIsRule(operator='between', formula=['1','5'], stopIfTrue=True, fill=redFill))

    # Format using a formula
    ws.conditional_formatting.add('E1:E10', FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))

    # Aside from the 2-color and 3-color scales, format rules take fonts, borders and fills for styling:
    myFont = Font()
    myBorder = Border()
    ws.conditional_formatting.add('E1:E10', FormulaRule(formula=['E1=0'], font=myFont, border=myBorder, fill=redFill))

    # Highlight cells that contain particular text by using a special formula
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws.conditional_formatting.add('A1:F40', rule)
    wb.save("test.xlsx")

#formatting_xls()

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    #print(rows)
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill