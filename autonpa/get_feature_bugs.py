from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from time import sleep
import pytest
from openpyxl import load_workbook as lw
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
#from openpyxl.formatting.rule import IconSet, FormatObject
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from test_pyxl_lib import style_range
import datetime


@pytest.fixture
def driver(request):
    chrome_options = Options()
    chrome_options.add_argument("--window-size=1920,1080")
    wd = webdriver.Chrome(chrome_options=chrome_options)

    print(wd.capabilities)
    request.addfinalizer(wd.quit)
    return wd

def cr_file_xls(filename):
   wb = Workbook()
   ws1=wb.active
   ws1.title = "Список"
   wb.save(filename = filename)
   return [filename, ws1.title]

def login(driver):
    with open('login_data.txt', 'r', encoding='utf-8') as ld:
        strings = ld.readlines()
        #print(strings)
        login_data = strings[0].split(',')
        proj_data = []
        for project in range(1, len(strings)):
            print(range(1, len(strings)), len(strings))
            proj_data.append(strings[project].split(','))

        #print(strings)
        #print(login_data)
    #login
    driver.get('http://jira.it2g.ru/login.jsp')
    driver.find_element_by_id('login-form-username').send_keys(login_data[0])
    driver.find_element_by_id('login-form-password').send_keys(login_data[1])
    driver.find_element_by_id('login-form-submit').click()
    #print(proj_data)
    return proj_data

# Функция возвращает массив задач по фильтру filter_id
def get_tasks_list(driver, filter_id, sprint, mode = 'proj_status'):
    # Открываем фильтр по его айди
    driver.get(f'http://jira.it2g.ru/issues/?filter={filter_id}')
    wait = WebDriverWait(driver, 5)
    quantity_of_tasks, tasks_on_page = 0,0
    try:
        wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'issue-link')))
        quantity_of_tasks = int(driver.find_element_by_class_name('results-count-total').text)
        tasks_on_page = int(driver.find_element_by_class_name('results-count-end').text)
    except:

        try:
            wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'issue-link')))
            quantity_of_tasks = int(driver.find_element_by_class_name('results-count-total').text)
            tasks_on_page = int(driver.find_element_by_class_name('results-count-end').text)
        except:
            print("Нет задач")
            return ''


    #print(str(tasks_on_page) + ' из ' + str(quantity_of_tasks))

    correct_iss, temp_iss = [], []
    while (quantity_of_tasks >= tasks_on_page):
        wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'issue-link')))
        issues = driver.find_elements_by_class_name('issue-link')
        try:
            issues = [x.text for x in issues]
        except:
            issues = driver.find_elements_by_class_name('issue-link')
            issues = [x.text for x in issues]
        issues = list(filter(lambda x: x!='', issues))
        issues = [issues[x] for x in range(0, len(issues), 2)]
        temp_iss += issues
        try:
            driver.find_element_by_class_name('icon-next').click()
        except:
            tasks_on_page+=quantity_of_tasks
            print('последняя страница')

    if mode == 'proj_status':
        correct_iss = temp_iss
        return correct_iss
        # versions = driver.find_elements_by_class_name('fixVersions')
        # versions = [x.text for x in versions]
        # # Вычислить номер следующего релиза
        # next_release = sprint.split(',')
        # # correct_sprints = []
        # # correct_sprints = [correct_sprints.append(x) for next_release[x] in range(1, len(next_release))]
        # next_release = next_release[0]
        # next_r_num = next_release[-1]
        # print(next_release)
        # next_release = 'Release ' + str(int(next_r_num) + 1)
        # print("Следующий релиз: " + next_release)
        # # Проверка версии у задачи перед записью в файл
        # for y in range(len(temp_iss)):
        #     if next_release not in versions[y]: #or ('Sprint 2' not in versions[y] or 'Sprint 3' not in versions[y]):
        #         correct_iss.append(temp_iss[y])
        #     else:
        #         print(f'Задача с релизом {next_release}: ' + temp_iss[y])
        # return correct_iss
    else:
        # print(correct_iss)
        correct_iss = temp_iss
        return correct_iss
    #print(issues, versions)

# Масссив для быстрой отладки скрипта для вкладки тестирование
# test_arr = ['NPA-1219', 'NPA-1429']

# Отсюда вызываем все функции этого модуля
def test_main(driver):
    fn = cr_file_xls("Тестовая таблица.xlsx")
    login(driver)
    tsk_list = []
    counter = 0
    tsk_list = get_tasks_list(driver, '10472')
    for u in range(len(tsk_list)):
        try:
            bgs = search_data(driver, tsk_list[u])
        except:
            print('Связанных багов нет.')
            bgs = []
        iss = tsk_data(driver, tsk_list[u])
        counter = write_to_xls(iss, bgs, fn, counter)


def search_data(driver, tasknum):

    # Открываем фильтр по багам связанным с задачей.
    driver.get('http://jira.it2g.ru/issues/?filter=10480')
    sleep(2)
    driver.find_element_by_id('advanced-search').clear()
    new_val=f'project = NPA AND issuetype = Bug AND issue in linkedIssues({str(tasknum)}) AND status != Закрыт'
    driver.find_element_by_id('advanced-search').send_keys(new_val)
    driver.find_elements_by_class_name('aui-icon')[8].click()
    sleep(2)
    return get_bugs_data(driver)
    
def get_bugs_data(driver):
    bugs, bugnums, summ =[], [], []

    bugs = driver.find_elements_by_class_name('issue-link')
    bugs = [x.text for x in bugs]
    bugs = list(filter(lambda x: x!='', bugs))

    # Разделяем номер бага и тему бага
    summ = [bugs[x] for x in range(1, len(bugs), 2)]
    bugnums = [bugs[x] for x in range(0, len(bugs), 2)]

    # Статус бага
    stat = driver.find_elements_by_class_name('status')
    stat = [x.text for x in stat]

    # Приоритет бага
    pr = driver.find_elements_by_xpath('//*[@class="priority"]/img')
    pr = [x.get_attribute('alt') for x in pr]

    # Исполнитель бага
    assign = driver.find_elements_by_class_name('assignee')
    assign = [x.text for x in assign]
    
    # Объединить данные в один массив
    result = [[] for x in range(len(bugnums))]
    # print(len(bugnums))
    
    for y in range(len(bugnums)):
        result[y].append(bugnums[y])
        result[y].append(summ[y])
        result[y].append(stat[y])
        result[y].append(pr[y])
        result[y].append(assign[y])

    return result

# Запись в файл эксель данных по задаче и по связанным багам
def write_to_xls(task, bugs, df, lr=0):
    headers = [
        '№ задачи Jira', 
        'Приоритет',
        'Статус',
        'Тема:',
        'Тестировщик',
        'Версия',
        'Комментарий', 
        '№ бага', 
        'Тема',  
        'Статус', 
        'Приоритет', 
        'Исполнитель', 
        'Можно перенести']


    # assign the icon set to a rule
    border = Border(left = Side(border_style = 'double', color = 'FF000000'),
                                          right = Side(border_style = 'double', color = 'FF000000'),
                                          top = Side(border_style = 'double', color = 'FF000000'),
                                          bottom = Side(border_style = 'double', color = 'FF000000'))

    wb = lw(df)
    ws1 = wb["В тестировании"]
    
    quan_h = ['Кол-во багов: ', str(len(bugs))]
    #Есть задача, баги, файл назначения и номер строки (который передаем)
    starts, ends = '', ''
    for row in range(lr+1, lr+2):
        # Запишем заголовки в файл
        starts = f'{get_column_letter(2)}{lr+1}'
        for col in range(0, len(headers)):
            _ = ws1.cell(column=col+2, row=row, value=headers[col])

        # Записываем данные по задаче
        print(len(headers), len(task))
        for col in range(0, len(headers)-7):
            _ = ws1.cell(column=col+2, row=row+1, value=task[col])

        for col in range(0, len(headers)-8):
            if(col != len(headers)-8):
                if bugs == []:
                    lr = row+4
                else:
                   # Записываем данные по багам
                   #print(len(headers), len(bugs[1]))
                   for b in range(1, len(bugs)+1):
                        #print(col+8)
                        _ = ws1.cell(column=col+9, row=row+b, value = bugs[b-1][col])
        ends = f'{get_column_letter(len(headers)+1)}{row+(len(bugs)+1)}'
        #print(f'Финальная ячейка: {ends}')
        lr = row+len(bugs)+2
        for col in range(0, len(quan_h)):
            # Кол-во багов под данными по задаче
            _ = ws1.cell(column=col+2, row=row+2, value = quan_h[col])
        print(task[0], quan_h)
    rang = f'{starts}:{ends}'
    print(rang)
    style_range(ws1, rang, border=border)
    #ws1.conditional_formatting.add(rang, border = dxf)
    wb.save(filename = df)
    return lr

def write_quantity_of_task(df, cntr):
    wb = lw(df)
    ws1 = wb["В тестировании"]
    quantity = ['Кол-во задач: ', '=СЧЁТЗ(E1:E99)/2']
    for col in range(0, len(quantity)):
        # Кол-во багов под данными по задаче
        _ = ws1.cell(column=col+2, row=cntr+1, value = quantity[col])
    wb.save(filename = df)

# Собираем данные о задаче возвращаем №, приоритет, статус, тему, QA
def tsk_data(driver, issue):
    task_res = []
    print(issue)
    sleep(3)
    driver.get(f'http://jira.it2g.ru/browse/{str(issue)}')

    task_res.append(issue)
    task_res.append(str(driver.find_element_by_id('priority-val').text))
    task_res.append(str(driver.find_element_by_id('status-val').text))
    task_res.append(str(driver.find_element_by_id('summary-val').text))
    task_res.append(str(driver.find_element_by_xpath('//*[@id="customfield_10201-val"]').text))
    task_res.append(str(driver.find_element_by_id('fixfor-val').text))
    print(task_res)
    return task_res

